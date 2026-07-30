#!/usr/bin/env python3
"""Importa campeões das planilhas do NCOFDI para data/competicoes/.

Uso:  python3 ferramentas/importar.py               # simulação de tudo
      python3 ferramentas/importar.py --aplicar     # grava
      python3 ferramentas/importar.py Europeus      # só um arquivo

Fonte primária: a coluna "Campeões" de cada aba, que lista os campeões na
ordem dos blocos de competição. O chaveamento é lido em paralelo e serve de
conferência — divergências são reportadas, não corrigidas em silêncio.
"""
import openpyxl, glob, os, re, io, json, sys, collections, unicodedata

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "ferramentas"))
import mapa_planilhas as M

PLACAR = re.compile(r"^\s*(\d+)\s*x\s*(\d+)\s*(?:\(\s*(\d+)\s*x\s*(\d+)\s*\))?\s*$", re.I)
NOME_ARQ = re.compile(r"Campeonatos (.+) - NCOFDI \(ANO (\d+)\)\.xlsx$")


def norm(t):
    t = unicodedata.normalize("NFD", str(t))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn").lower()
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


NAO_COMP = {norm(x) for x in M.NAO_COMPETICAO}
SUBS = {norm(x) for x in M.SUB_TORNEIOS}
DECIS = {norm(x) for x in M.DECISIVOS}


# ------------------------------------------------------------------ planilha
def blocos(ws, arq_base, aba):
    """Blocos de competição da aba, em ordem, como (coluna, nome)."""
    achados = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v not in (None, ""):
            achados.append((c, str(v).strip()))
    for extra in M.BLOCOS_LINHA_2.get((arq_base, aba), []):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(2, c).value).strip() == extra:
                achados.append((c, extra))
    achados.sort()
    return [(c, n) for c, n in achados
            if norm(n) not in NAO_COMP and not n.startswith("*")]


def coluna_campeoes(ws):
    col = next((c for c in range(1, ws.max_column + 1)
                if norm(ws.cell(1, c).value) in ("campeoes", "campeao")), None)
    if not col:
        return None
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v not in (None, "") and str(v).strip():
            out.append(str(v).strip())
        elif out:
            break
    return out


def ler_placar(t):
    m = PLACAR.match(str(t or ""))
    return None if not m else (int(m.group(1)), int(m.group(2)),
                               int(m.group(3)) if m.group(3) else None,
                               int(m.group(4)) if m.group(4) else None)


def vencedor(pernas):
    gols, pen = collections.Counter(), None
    for casa, placar, fora in pernas:
        r = ler_placar(placar)
        if r is None:
            return None
        g1, g2, p1, p2 = r
        gols[casa] += g1
        gols[fora] += g2
        if p1 is not None:
            pen = (casa, p1, fora, p2)
    if len(gols) != 2:
        return None
    (t1, a), (t2, b) = gols.items()
    if a != b:
        return t1 if a > b else t2
    if pen and pen[1] != pen[3]:
        return pen[0] if pen[1] > pen[3] else pen[2]
    return None


def confronto(ws, linha, col):
    a, p, b = (ws.cell(linha, col).value, ws.cell(linha, col + 1).value,
               ws.cell(linha, col + 2).value)
    if not (a and p and b):
        return []
    pernas = [(str(a).strip(), p, str(b).strip())]
    a2, p2, b2 = (ws.cell(linha, col + 4).value, ws.cell(linha, col + 5).value,
                  ws.cell(linha, col + 6).value)
    if a2 and p2 and b2:
        pernas.append((str(a2).strip(), p2, str(b2).strip()))
    return pernas


def campeao_por_chaveamento(ws, col):
    """Campeão do bloco lendo a final (ou a classificação, se não houver)."""
    finais, geral, corte = [], None, None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str):
            continue
        t = norm(v)
        if t == "final":
            finais.append(r)
        elif t in DECIS:
            geral = r
        elif t in SUBS and corte is None:
            corte = r
    if geral:
        return vencedor(confronto(ws, geral + 1, col))
    validas = [r for r in finais if corte is None or r < corte]
    if validas:
        return vencedor(confronto(ws, validas[0], col)) or \
               vencedor(confronto(ws, validas[0] + 1, col))
    tabelas = [r for r in range(2, ws.max_row + 1)
               if str(ws.cell(r, col).value).strip() == "#"]
    if tabelas:
        v = ws.cell(tabelas[0] + 1, col + 1).value
        return str(v).strip() if v else None
    return None


def celulas_campeao(ws):
    """Valores logo abaixo da célula rotulada 'Campeão'/'Campeões'."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(r, c).value) in ("campeao", "campeoes"):
                out = []
                for rr in range(r + 1, ws.max_row + 1):
                    v = ws.cell(rr, c).value
                    if v not in (None, "") and str(v).strip():
                        out.append(str(v).strip())
                    elif out:
                        break
                return out
    return []


def campeao_pela_ultima_final(ws):
    """Reserva para abas sem célula 'Campeão': lê a final mais adiantada da aba."""
    alvo = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(r, c).value) == "final":
                alvo = (r, c)
    if not alvo:
        return None
    r, c = alvo
    return vencedor(confronto(ws, r + 1, c))


def lider_da_tabela(ws, nome_bloco):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).strip() == nome_bloco:
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, c).value).strip() == "#":
                    v = ws.cell(r + 1, c + 1).value
                    return str(v).strip() if v else None
    return None



def resolver(bloco, i, aba, nomes_site):
    """Nome da competição do site correspondente ao bloco `bloco` (i-ésimo da aba)."""
    apelido = M.ALIASES.get((aba, bloco))
    if apelido:
        return apelido
    alvo = norm(bloco)
    for n in nomes_site:                       # nome idêntico
        if norm(n) == alvo:
            return n
    if alvo in ("apertura", "clausura"):       # 'Apertura' -> 'Liga MX — Apertura'
        for n in nomes_site:
            if norm(n).endswith(" " + alvo):
                return n
    if i < len(nomes_site):                    # posição (estaduais e afins)
        return nomes_site[i]
    return bloco


# ---------------------------------------------------------------------- site
def carregar_site():
    est = json.load(io.open(os.path.join(RAIZ, "data", "estrutura.json"), encoding="utf-8"))
    locais, por_nome = {}, {}
    for s in est["secoes"]:
        for x in (s.get("confederacoes") or []) + (s.get("paises") or []):
            cam = os.path.join(RAIZ, "data", "competicoes", x["id"] + ".json")
            d = json.load(io.open(cam, encoding="utf-8"))
            locais[x["id"]] = (cam, d, {norm(c["nome"]): c["id"] for c in d["competicoes"]})
            por_nome[norm(x["nome"])] = x["id"]
            if x.get("sigla"):
                por_nome[norm("br-" + x["sigla"])] = x["id"]
    return locais, por_nome


def carregar_sufixos():
    comuns = set()
    with io.open(os.path.join(RAIZ, "ferramentas", "nomes-comuns.txt"), encoding="utf-8") as f:
        for l in f:
            l = l.split("#")[0].strip()
            if l:
                comuns.add(norm(l))
    siglas = {}
    with io.open(os.path.join(RAIZ, "ferramentas", "siglas-paises.txt"), encoding="utf-8") as f:
        for l in f:
            if "\t" in l and not l.startswith("#"):
                pais, sig = l.split("\t")
                siglas[norm(pais)] = sig.strip()
    return comuns, siglas


# -------------------------------------------------------------------- import
def main():
    alvo = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
    aplicar = "--aplicar" in sys.argv

    locais, local_por_nome = carregar_site()
    comuns, siglas = carregar_sufixos()

    registros, avisos, divergencias = [], [], []

    for caminho in sorted(glob.glob(os.path.join(RAIZ, "fontes", "*.xlsx"))):
        m = NOME_ARQ.search(os.path.basename(caminho))
        if not m:
            continue
        base, ano_n = m.group(1), int(m.group(2))
        if alvo and alvo.lower() not in base.lower():
            continue
        if base == "Brasileiros":
            for aba in openpyxl.load_workbook(caminho, data_only=True).sheetnames:
                ws = openpyxl.load_workbook(caminho, data_only=True)[aba]
                if aba not in M.BRASILEIROS:
                    avisos.append(f"Brasileiros/{aba}: aba desconhecida")
                    continue
                lid, nomes = M.BRASILEIROS[aba]
                _cam, dados, comp_por_nome = locais[lid]
                achados = celulas_campeao(ws)
                if not achados and len(nomes) == 1:
                    unico = campeao_pela_ultima_final(ws)
                    if unico:
                        achados = [unico]
                        avisos.append(f"Brasileiros/{aba} {M.temporada(ano_n)}: sem célula "
                                      f"'Campeão'; li a final: {unico}")
                if len(achados) != len(nomes):
                    avisos.append(f"Brasileiros/{aba} {M.temporada(ano_n)}: "
                                  f"{len(achados)} campeões para {len(nomes)} competições — PULADO")
                    continue
                for nome_comp, clube in zip(nomes, achados):
                    cid = comp_por_nome.get(norm(nome_comp))
                    if cid:
                        registros.append((lid, cid, nome_comp, clube, base, aba,
                                          M.temporada(ano_n), "celula"))
                    else:
                        avisos.append(f"Brasileiros/{aba}: '{nome_comp}' não existe em {lid}.json")
            continue
        temp = M.temporada(ano_n)
        wb = openpyxl.load_workbook(caminho, data_only=True)

        for aba in wb.sheetnames:
            ws = wb[aba]
            lid = M.ABA_PARA_LOCAL.get((base, aba)) or local_por_nome.get(norm(aba))
            if base == "Estaduais":
                lid = f"br-{aba.lower()}"
            if lid not in locais:
                avisos.append(f"{base}/{aba}: local '{lid}' não existe no site")
                continue
            _cam, dados, comp_por_nome = locais[lid]

            bl = blocos(ws, base, aba)
            lista = coluna_campeoes(ws)

            # ---- quais competições do site esta aba alimenta, em ordem ----
            chave = M.EXPLICITO.get((base, aba, temp)) or M.EXPLICITO.get((base, aba, None))
            if chave:
                alvos = chave
            else:
                nomes_site = [c["nome"] for c in dados["competicoes"]]
                alvos = [resolver(n, i, aba, nomes_site) for i, (_c, n) in enumerate(bl)]

            # ---- campeões ----
            if (base, aba, temp) in M.POR_CHAVEAMENTO or lista is None:
                achados = [campeao_por_chaveamento(ws, c) for c, _n in bl]
                origem = "chaveamento"
            else:
                achados = list(lista)
                origem = "coluna"

            if len(achados) != len(alvos):
                avisos.append(f"{base}/{aba} {temp}: {len(achados)} campeões para "
                              f"{len(alvos)} competições — PULADO")
                continue

            for nome_comp, clube in zip(alvos, achados):
                if not clube:
                    avisos.append(f"{base}/{aba} {temp}: sem campeão para '{nome_comp}'")
                    continue
                cid = comp_por_nome.get(norm(nome_comp))
                if not cid:
                    avisos.append(f"{base}/{aba} {temp}: competição '{nome_comp}' "
                                  f"não existe em {lid}.json")
                    continue
                registros.append((lid, cid, nome_comp, clube, base, aba, temp, origem))

            # ---- complementos (campeão fora da coluna) ----
            for nome_comp, nome_bloco in M.COMPLEMENTOS.get((base, aba, temp), []):
                clube = lider_da_tabela(ws, nome_bloco)
                cid = comp_por_nome.get(norm(nome_comp))
                if clube and cid:
                    registros.append((lid, cid, nome_comp, clube, base, aba, temp, "tabela"))
                    avisos.append(f"{base}/{aba} {temp}: '{nome_comp}' veio da classificação "
                                  f"(não estava na coluna): {clube}")

            # ---- conferência: coluna vs chaveamento ----
            if origem == "coluna" and not chave:
                for (c, nb), decl in zip(bl, achados):
                    calc = campeao_por_chaveamento(ws, c)
                    if calc and norm(calc) != norm(decl):
                        divergencias.append(f"{base}/{aba} {temp} · {nb}: "
                                            f"coluna='{decl}' chaveamento='{calc}'")

    # ------------------------------------------------------------ sufixos
    cam_idx = os.path.join(RAIZ, "ferramentas", "indice-clubes.json")
    indice = json.load(io.open(cam_idx, encoding="utf-8")) if os.path.exists(cam_idx) else None
    nao_resolvidos = []

    def sufixo(lid, clube):
        """Nome canônico do clube.

        Quando o local já determina a origem (estado ou país), ele manda: é a
        informação mais específica que existe. O índice das planilhas só entra
        quando o local não diz nada — confederações e competições nacionais
        brasileiras, onde o clube pode ser de qualquer estado.
        """
        # clube que mudou de nome: a planilha ainda traz o antigo
        pais_local = locais[lid][1]["nome"]
        clube = getattr(M, "RENOMEADOS", {}).get((pais_local, clube), clube)
        comum = norm(clube) in comuns

        if lid.startswith("br-"):                       # estadual
            return f"{clube}-{lid[3:].upper()}" if comum else clube
        sig = siglas.get(norm(locais[lid][1]["nome"]))
        if sig:                                         # país
            return f"{clube}-{sig}" if comum else clube

        # o local não decide: vale o mapeamento manual, depois o índice
        if clube in M.CLUBES_MANUAIS:
            return M.CLUBES_MANUAIS[clube][0]
        if indice:                                      # confederação ou nacional brasileira
            if norm(clube) in indice["final"]:
                return indice["final"][norm(clube)]
            achado = indice["base"].get(norm(clube))
            if achado:
                return achado
        if comum:
            nao_resolvidos.append((lid, clube))
        return clube

    finais = [(lid, cid, nc, sufixo(lid, cl), base, aba, temp, org)
              for lid, cid, nc, cl, base, aba, temp, org in registros]

    # -------------------------------------------------------------- saída
    por_temp = collections.Counter(r[6] for r in finais)
    print(f"{len(finais)} campeões · " + " · ".join(f"{t}: {n}" for t, n in sorted(por_temp.items())))
    print(f"origem: " + " · ".join(f"{k}={v}" for k, v in
                                   collections.Counter(r[7] for r in finais).items()))

    suf = [r for r in finais if r[3] != next(x[3] for x in registros
                                             if (x[0], x[1], x[6]) == (r[0], r[1], r[6]))]
    print(f"\n{len(suf)} nomes receberam sufixo")

    if divergencias:
        print(f"\n── Coluna 'Campeões' × chaveamento: {len(divergencias)} divergências ──")
        for d in divergencias:
            print("   ⚠", d)
    else:
        print("\n✓ coluna e chaveamento concordam em todos os blocos conferíveis")

    if nao_resolvidos:
        print(f"\n⚠ {len(nao_resolvidos)} campeões de nome comum sem entrada única no índice: "
              + ", ".join(f"{c} ({l})" for l, c in sorted(set(nao_resolvidos))))

    if avisos:
        print(f"\n── Avisos ({len(avisos)}) ──")
        for a in avisos:
            print("   •", a)

    if not aplicar:
        print("\n[simulação] rode com --aplicar para gravar")
        return

    for lid, nome_comp, temp, clube in getattr(M, "CAMPEOES_MANUAIS", []):
        cid = locais[lid][2].get(norm(nome_comp))
        if cid:
            finais.append((lid, cid, nome_comp, clube, "manual", "-", temp, "manual"))
            print(f"   + {clube} — {nome_comp} {temp} (dos slides, ausente na planilha)")

    porarq = collections.defaultdict(list)
    for lid, cid, _nc, clube, _b, _a, temp, _o in finais:
        porarq[lid].append((cid, temp, clube))
    for lid, itens in porarq.items():
        cam, dados, _c = locais[lid]
        for cid, temp, clube in itens:
            comp = next(c for c in dados["competicoes"] if c["id"] == cid)
            comp["campeoes"] = [t for t in comp.get("campeoes", [])
                                if str(t.get("temporada")) != str(temp)]
            comp["campeoes"].append({"temporada": str(temp), "clube": clube})
            comp["campeoes"].sort(key=lambda t: str(t["temporada"]))
        json.dump(dados, io.open(cam, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        io.open(cam, "a", encoding="utf-8").write("\n")
    print(f"\n✓ gravado em {len(porarq)} arquivos")


if __name__ == "__main__":
    main()
