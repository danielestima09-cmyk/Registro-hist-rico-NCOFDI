#!/usr/bin/env python3
"""Importa os campeões dos estaduais a partir da planilha do NCOFDI.

Uso:  python3 ferramentas/importar_estaduais.py "fontes/Campeonatos Estaduais - NCOFDI (ANO 1).xlsx" 2026
      python3 ferramentas/importar_estaduais.py <planilha> <temporada> --aplicar

Sem --aplicar, só mostra o que faria (simulação).

A planilha tem uma aba por estado. Cada aba tem blocos de competição na linha 1
e, dentro de cada bloco, a fase "Final" com o confronto decisivo.
"""
import openpyxl, re, sys, json, io, os, collections, unicodedata

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LISTA_NOMES = os.path.join(RAIZ, "ferramentas", "nomes-comuns.txt")

# Blocos da linha 1 que não são competições, e sim quadros de apoio.
# Comparados sem acento e em minúsculas (ver normalizar() abaixo).
NAO_COMPETICAO_BRUTO = ["Tabela geral", "Campeão", "Campeões"]

# Troféus secundários dentro do bloco do estadual: têm final própria, mas o
# usuário não os registra. A final válida é a que vem ANTES deles.
SECUNDARIOS = {"taça rio", "taça farroupilha", "taça acesc", "troféu inconfidência"}

# Confrontos que a planilha não resolve sozinha (regulamento não fica no arquivo).
DESEMPATES = {
    ("AL", "Alagoano B"): ("Jaciobá",
        "vantagem do time de melhor campanha na 1ª fase: empata no agregado e avança"),
}

# Grafias divergentes do mesmo clube dentro da planilha.
PADRONIZAR = {
    ("SC", "Internacional"): "Inter de Lages",
}

PLACAR = re.compile(r"^\s*(\d+)\s*x\s*(\d+)\s*(?:\(\s*(\d+)\s*x\s*(\d+)\s*\))?\s*$", re.I)


def sem_acento(txt):
    t = unicodedata.normalize("NFD", str(txt))
    return "".join(c for c in t if unicodedata.category(c) != "Mn").lower().strip()


NAO_COMPETICAO = {sem_acento(x) for x in NAO_COMPETICAO_BRUTO}
SECUNDARIOS_NORM = {sem_acento(x) for x in SECUNDARIOS}


def carregar_nomes_comuns():
    nomes = set()
    with io.open(LISTA_NOMES, encoding="utf-8") as f:
        for linha in f:
            linha = linha.split("#")[0].strip()
            if linha:
                nomes.add(sem_acento(linha))
    return nomes


def ler_placar(txt):
    m = PLACAR.match(str(txt or ""))
    if not m:
        return None
    return (int(m.group(1)), int(m.group(2)),
            int(m.group(3)) if m.group(3) else None,
            int(m.group(4)) if m.group(4) else None)


def vencedor(pernas):
    """pernas = [(casa, placar, fora), ...] -> (vencedor, erro)"""
    gols, penaltis = collections.Counter(), None
    for casa, placar, fora in pernas:
        r = ler_placar(placar)
        if r is None:
            return None, f"placar ilegível: {placar!r}"
        g1, g2, p1, p2 = r
        gols[casa] += g1
        gols[fora] += g2
        if p1 is not None:
            penaltis = (casa, p1, fora, p2)
    if len(gols) != 2:
        return None, f"confronto com {len(gols)} times"
    (t1, a), (t2, b) = gols.items()
    if a != b:
        return (t1 if a > b else t2), None
    if penaltis:
        c, pc, f, pf = penaltis
        if pc != pf:
            return (c if pc > pf else f), None
    return None, f"agregado empatado em {a}x{b} sem desempate registrado"


def confronto_em(ws, linha, col):
    """Lê o confronto da linha: ida em col..col+2, volta (se houver) em col+4..col+6."""
    casa, placar, fora = (ws.cell(linha, col).value,
                          ws.cell(linha, col + 1).value,
                          ws.cell(linha, col + 2).value)
    if not (casa and placar and fora):
        return []
    pernas = [(str(casa).strip(), placar, str(fora).strip())]
    c2, p2, f2 = (ws.cell(linha, col + 4).value,
                  ws.cell(linha, col + 5).value,
                  ws.cell(linha, col + 6).value)
    if c2 and p2 and f2:
        pernas.append((str(c2).strip(), p2, str(f2).strip()))
    return pernas


def blocos(ws):
    marcas = [(c, str(ws.cell(1, c).value).strip())
              for c in range(1, ws.max_column + 1)
              if ws.cell(1, c).value not in (None, "")]
    return [(c, n) for c, n in marcas if sem_acento(n) not in NAO_COMPETICAO]


def campeao(ws, col):
    """(campeao, metodo, erro) para o bloco que começa na coluna col."""
    finais, geral, secundario = [], None, None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str):
            continue
        t = sem_acento(v)
        if t == "final":
            finais.append(r)
        elif t == "final geral":
            geral = r
        elif t in SECUNDARIOS_NORM:
            secundario = r

    if geral:                                   # turno/returno: decide a final geral
        p = confronto_em(ws, geral + 1, col)
        v, e = vencedor(p) if p else (None, "sem confronto após 'Final geral'")
        return v, "final geral", e

    if finais:                                  # com troféu secundário, vale a final anterior a ele
        validas = [r for r in finais if not secundario or r < secundario]
        alvo = (validas or finais)[0]
        p = confronto_em(ws, alvo + 1, col)
        v, e = vencedor(p) if p else (None, "sem confronto após 'Final'")
        return v, "final", e

    tabelas = [r for r in range(2, ws.max_row + 1)
               if str(ws.cell(r, col).value).strip() == "#"]
    if not tabelas:
        return None, "?", "bloco sem final e sem classificação"
    lider = ws.cell(tabelas[-1] + 1, col + 1).value
    return (str(lider).strip() if lider else None), "classificação", None


def com_sufixo(nome, uf, comuns):
    return f"{nome}-{uf}" if sem_acento(nome) in comuns else nome


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    planilha, temporada = sys.argv[1], sys.argv[2]
    aplicar = "--aplicar" in sys.argv

    comuns = carregar_nomes_comuns()
    wb = openpyxl.load_workbook(planilha, data_only=True)

    registros, avisos, pulados = [], [], []
    for uf in wb.sheetnames:
        ws = wb[uf]
        arquivo = os.path.join(RAIZ, "data", "competicoes", f"br-{uf.lower()}.json")
        if not os.path.exists(arquivo):
            avisos.append(f"{uf}: não existe data/competicoes/br-{uf.lower()}.json")
            continue
        site = json.load(io.open(arquivo, encoding="utf-8"))

        for i, (col, nome_bloco) in enumerate(blocos(ws)):
            if i >= len(site["competicoes"]):
                avisos.append(f"{uf}: bloco '{nome_bloco}' não tem competição correspondente no site")
                continue
            comp = site["competicoes"][i]

            clube, metodo, erro = campeao(ws, col)
            chave = (uf, nome_bloco)
            if chave in DESEMPATES:
                clube, motivo = DESEMPATES[chave]
                metodo, erro = "regulamento", None
                avisos.append(f"{uf} / {comp['nome']}: {clube} — {motivo}")
            if clube and (uf, clube) in PADRONIZAR:
                novo = PADRONIZAR[(uf, clube)]
                avisos.append(f"{uf} / {comp['nome']}: grafia '{clube}' padronizada para '{novo}'")
                clube = novo

            if not clube:
                avisos.append(f"{uf} / {comp['nome']}: SEM CAMPEÃO — {erro}")
                continue
            registros.append((uf, arquivo, comp["id"], comp["nome"],
                              com_sufixo(clube, uf, comuns), metodo))

        # competições do site sem bloco na planilha (não disputadas neste ano)
        for comp in site["competicoes"][len(blocos(ws)):]:
            pulados.append(f"{uf} / {comp['nome']}")

    # ---------------------------------------------------------------- saída
    print(f"{'UF':<4}{'competição':<44}{'campeão':<26}{'como'}")
    print("-" * 98)
    for uf, _arq, _cid, cnome, clube, metodo in registros:
        print(f"{uf:<4}{cnome:<44}{clube:<26}{metodo}")
    print("-" * 98)
    print(f"{len(registros)} campeões · temporada {temporada}")

    com_suf = [r for r in registros if r[4].endswith(f"-{r[0]}")]
    print(f"\n{len(com_suf)} nomes receberam sufixo de estado:")
    for uf, _a, _c, cnome, clube, _m in com_suf:
        print(f"   {clube:<26} ({cnome})")

    if pulados:
        print(f"\n{len(pulados)} competições do site sem bloco na planilha (ficam sem campeão):")
        for p in pulados:
            print("   ·", p)

    if avisos:
        print(f"\nAvisos ({len(avisos)}):")
        for a in avisos:
            print("   •", a)

    if not aplicar:
        print("\n[simulação] rode de novo com --aplicar para gravar em data/competicoes/")
        return

    por_arquivo = collections.defaultdict(list)
    for _uf, arq, cid, _cn, clube, _m in registros:
        por_arquivo[arq].append((cid, clube))

    for arq, itens in por_arquivo.items():
        d = json.load(io.open(arq, encoding="utf-8"))
        for cid, clube in itens:
            comp = next(c for c in d["competicoes"] if c["id"] == cid)
            # idempotente: substitui a entrada desta temporada, se já existir
            comp["campeoes"] = [t for t in comp.get("campeoes", [])
                                if str(t.get("temporada")) != str(temporada)]
            comp["campeoes"].append({"temporada": temporada, "clube": clube})
            comp["campeoes"].sort(key=lambda t: str(t["temporada"]))
        json.dump(d, io.open(arq, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        io.open(arq, "a", encoding="utf-8").write("\n")

    print(f"\n✓ gravado em {len(por_arquivo)} arquivos")


if __name__ == "__main__":
    main()
