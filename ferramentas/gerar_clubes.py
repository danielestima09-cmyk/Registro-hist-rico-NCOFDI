#!/usr/bin/env python3
"""Monta data/clubes.json a partir das planilhas.

Cada aba de país (ou estado) lista, nas tabelas de classificação e nos
confrontos, todos os clubes daquele país. Isso dá o país de origem de cada
clube — inclusive dos que só aparecem no site como campeões continentais,
onde o site não teria como deduzir sozinho.

Uso:  python3 ferramentas/gerar_clubes.py [--aplicar]
"""
import openpyxl, glob, os, re, io, json, sys, collections, unicodedata

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "ferramentas"))
import mapa_planilhas as M

NOME_ARQ = re.compile(r"Campeonatos (.+) - NCOFDI \(ANO (\d+)\)\.xlsx$")
PLACAR = re.compile(r"^\s*\d+\s*x\s*\d+\s*(?:\(\s*\d+\s*x\s*\d+\s*\))?\s*$", re.I)
POSICAO = re.compile(r"^\d+\.?$")

# Abas que reúnem clubes de vários países: não servem para inferir origem.
ABAS_MISTAS = {"Internacionais"}

# Cabeçalhos e rótulos que não são nome de clube.
RUIDO = {"time", "pts", "pj", "vit", "e", "der", "gp", "gc", "sg", "#", "campeoes",
         "campeao", "final", "semifinal", "final geral", "grupo a", "grupo b",
         "grupo c", "grupo d", "f", "tabela geral", "tabela acumulada"}


def norm(t):
    t = unicodedata.normalize("NFD", str(t))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn").lower()
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


def carregar_listas():
    comuns = set()
    with io.open(os.path.join(RAIZ, "ferramentas", "nomes-comuns.txt"), encoding="utf-8") as f:
        for l in f:
            l = l.split("#")[0].strip()
            if l:
                comuns.add(norm(l))
    siglas = {}
    with io.open(os.path.join(RAIZ, "ferramentas", "siglas-paises.txt"), encoding="utf-8") as f:
        for l in f:
            if "\t" in l and not l.startswith("#"):
                p, s = l.split("\t")
                siglas[norm(p)] = s.strip()
    return comuns, siglas


def clubes_da_aba(ws):
    """Todos os nomes de clube que aparecem na aba."""
    achados = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str) or not v.strip():
                continue
            txt = v.strip()
            # nome logo à direita de uma posição de tabela ("1." | "Santos")
            if POSICAO.match(txt):
                lado = ws.cell(r, c + 1).value
                if isinstance(lado, str) and lado.strip():
                    achados.add(lado.strip())
            # nomes dos dois lados de um placar ("Vasco" | "1x0" | "Flamengo")
            if PLACAR.match(txt):
                for dc in (-1, 1):
                    lado = ws.cell(r, c + dc).value
                    if isinstance(lado, str) and lado.strip():
                        achados.add(lado.strip())
    return {a for a in achados if norm(a) not in RUIDO and not a.startswith("*")}


def main():
    aplicar = "--aplicar" in sys.argv
    comuns, siglas = carregar_listas()

    est = json.load(io.open(os.path.join(RAIZ, "data", "estrutura.json"), encoding="utf-8"))
    pais_por_aba = {}
    for s in est["secoes"]:
        for p in (s.get("paises") or []):
            if p.get("sigla"):
                pais_por_aba[p["sigla"]] = ("Brasil", p["sigla"])   # estado -> país Brasil
            else:
                pais_por_aba[norm(p["nome"])] = (p["nome"], siglas.get(norm(p["nome"])))

    origem = collections.defaultdict(set)      # nome final -> {país}
    for caminho in sorted(glob.glob(os.path.join(RAIZ, "fontes", "*.xlsx"))):
        m = NOME_ARQ.search(os.path.basename(caminho))
        if not m or m.group(1) in ABAS_MISTAS:
            continue
        base = m.group(1)
        wb = openpyxl.load_workbook(caminho, data_only=True)
        for aba in wb.sheetnames:
            if base == "Brasileiros":
                pais, sufixo = "Brasil", None      # sufixo estadual não se aplica aqui
            elif base == "Estaduais":
                pais, sufixo = "Brasil", aba
            else:
                lid = M.ABA_PARA_LOCAL.get((base, aba))
                chave = norm(lid.replace("-", " ")) if lid else norm(aba)
                if chave not in pais_por_aba:
                    continue
                pais, sufixo = pais_por_aba[chave]
            for clube in clubes_da_aba(wb[aba]):
                nome = f"{clube}-{sufixo}" if (sufixo and norm(clube) in comuns) else clube
                origem[nome].add(pais)

    for _bruto, (canon, pais) in M.CLUBES_MANUAIS.items():
        origem[canon] = {pais}

    conflitos = {k: sorted(v) for k, v in origem.items() if len(v) > 1}
    clubes = {k: {"pais": next(iter(v))} for k, v in sorted(origem.items()) if len(v) == 1}

    print(f"{len(origem)} clubes encontrados nas planilhas")
    print(f"{len(clubes)} com país único · {len(conflitos)} em conflito")
    if conflitos:
        print("\nConflitos (mesmo nome em países diferentes) — ficam de fora:")
        for k, v in sorted(conflitos.items())[:25]:
            print(f"   {k}: {v}")

    # cobertura: campeões que continuam sem país
    faltando = []
    for s in est["secoes"]:
        for x in (s.get("confederacoes") or []) + (s.get("paises") or []):
            d = json.load(io.open(os.path.join(RAIZ, "data", "competicoes", x["id"] + ".json"),
                                  encoding="utf-8"))
            eh_conf = any(c["id"] == x["id"] for c in (s.get("confederacoes") or []))
            for c in d["competicoes"]:
                for t in c.get("campeoes", []):
                    if eh_conf and t["clube"] not in clubes:
                        faltando.append((x["nome"], c["nome"], t["temporada"], t["clube"]))
    print(f"\nCampeões continentais/mundiais ainda sem país: {len(faltando)}")
    for l, c, t, cl in faltando:
        print(f"   {cl}  ({c} {t})")

    # índice para o importador: nome-base -> nome final (quando não é ambíguo)
    base_para_final = collections.defaultdict(set)
    for final in clubes:
        base = re.sub(r"-[A-ZÀ-Ú]{2,3}$", "", final)
        base_para_final[norm(base)].add(final)
    indice = {
        "final": {norm(k): k for k in clubes},
        "base": {b: sorted(v)[0] for b, v in base_para_final.items() if len(v) == 1},
        "pais": {k: v["pais"] for k, v in clubes.items()},
    }

    if aplicar:
        json.dump(indice, io.open(os.path.join(RAIZ, "ferramentas", "indice-clubes.json"),
                                  "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        cam = os.path.join(RAIZ, "data", "clubes.json")
        json.dump(clubes, io.open(cam, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        io.open(cam, "a", encoding="utf-8").write("\n")
        print(f"\n✓ data/clubes.json gravado com {len(clubes)} clubes")
    else:
        print("\n[simulação] rode com --aplicar para gravar")


if __name__ == "__main__":
    main()
